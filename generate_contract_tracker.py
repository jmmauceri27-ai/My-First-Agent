from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from datetime import date
import calendar

wb = Workbook()
wb.calculation.calcMode = "auto"
wb.calculation.fullCalcOnLoad = True

# ── Palette ───────────────────────────────────────────────────────────────────
C_NAVY      = "1F3864"
C_BLUE      = "2E75B6"
C_WHITE     = "FFFFFF"
C_ALT       = "DCE6F1"
C_TODAY     = "FF0000"
C_BORDER    = "BDD7EE"
GANTT_COLORS = [
    "2E75B6", "70AD47", "ED7D31", "FFD966",
    "9DC3E6", "A9D18E", "FF7C80", "C5A3FF",
]

thin  = Side(style="thin",   color=C_BORDER)
thick = Side(style="medium", color=C_NAVY)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def border(top=thin, bottom=thin, left=thin, right=thin):
    return Border(top=top, bottom=bottom, left=left, right=right)

# ── Sample contracts (edit here or directly in Contract Data sheet) ───────────
contracts = [
    # (Client,             Trade/Service,          Start,              End)
    ("Acme Corp",          "Electrical",            date(2025,  1, 15), date(2025, 12, 31)),
    ("Acme Corp",          "HVAC",                  date(2025,  3,  1), date(2026,  2, 28)),
    ("Blue Ridge LLC",     "Plumbing",              date(2025,  2,  1), date(2025,  7, 31)),
    ("Blue Ridge LLC",     "General Contracting",   date(2025,  6,  1), date(2026,  5, 31)),
    ("Summit Partners",    "Roofing",               date(2025,  4,  1), date(2025,  9, 30)),
    ("Summit Partners",    "Electrical",            date(2025,  9,  1), date(2026,  8, 31)),
    ("Horizon Group",      "Landscaping",           date(2025,  1,  1), date(2025, 12, 31)),
    ("Horizon Group",      "Painting",              date(2025,  5, 15), date(2025, 11, 14)),
    ("Crestview Inc",      "Flooring",              date(2025,  7,  1), date(2026,  6, 30)),
    ("Crestview Inc",      "HVAC",                  date(2025, 10,  1), date(2026,  3, 31)),
]

# ── Timeline span ─────────────────────────────────────────────────────────────
tl_start = date(min(c[2] for c in contracts).year, 1, 1)
tl_end   = date(max(c[3] for c in contracts).year, 12, 31)

months = []
d = tl_start
while d <= tl_end:
    months.append((d.year, d.month))
    m = d.month + 1
    y = d.year + (1 if m > 12 else 0)
    d = date(y, 1 if m > 12 else m, 1)

today = date.today()

# ═══════════════════════════════════════════════════════════════════════════════
#  SHEET 1 — Contract Data  (source of truth)
# ═══════════════════════════════════════════════════════════════════════════════
ws_data = wb.active
ws_data.title = "Contract Data"

# Column layout for Contract Data
DC = {
    "num":     1,
    "client":  2,
    "trade":   3,
    "ctr_num": 4,
    "start":   5,
    "end":     6,
    "dur":     7,
    "renew":   8,
    "notice":  9,
    "value":  10,
    "notes":  11,
}
DC_WIDTHS = [5, 22, 22, 16, 13, 13, 16, 12, 20, 14, 30]
DC_HEADERS = ["#", "Client", "Trade / Service", "Contract #",
              "Start Date", "End Date", "Duration (days)",
              "Auto-Renew?", "Notice Period (days)", "Value ($)", "Notes"]

# Title row
ws_data.row_dimensions[1].height = 28
ws_data.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(DC_HEADERS))
t = ws_data.cell(1, 1, "CONTRACT MASTER LIST")
t.font      = Font(name="Calibri", bold=True, size=14, color=C_WHITE)
t.fill      = fill(C_NAVY)
t.alignment = Alignment(horizontal="center", vertical="center")

# Header row
ws_data.row_dimensions[2].height = 20
for ci, (h, w) in enumerate(zip(DC_HEADERS, DC_WIDTHS), 1):
    ws_data.column_dimensions[get_column_letter(ci)].width = w
    c = ws_data.cell(2, ci, h)
    c.font      = Font(name="Calibri", bold=True, size=10, color=C_WHITE)
    c.fill      = fill(C_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = border(top=thick, bottom=thick)

# Data rows — dates stored as real Excel date values
for ri, (client, trade, start, end) in enumerate(contracts, 1):
    r        = ri + 2
    row_fill = fill(C_ALT if ri % 2 == 0 else C_WHITE)
    dur_formula = f"=IF(AND(E{r}<>\"\",F{r}<>\"\"),F{r}-E{r}+1,\"\")"
    vals = [ri, client, trade, f"CTR-{ri:03d}", start, end, dur_formula, "No", 30, "", ""]
    for ci, val in enumerate(vals, 1):
        c = ws_data.cell(r, ci, val)
        c.fill      = row_fill
        c.font      = Font(name="Calibri", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = border()
    # dates: apply date format
    for col in (DC["start"], DC["end"]):
        ws_data.cell(r, col).number_format = "MM/DD/YYYY"
    # left-align text columns
    for col in (DC["client"], DC["trade"]):
        ws_data.cell(r, col).alignment = Alignment(horizontal="left", vertical="center", indent=1)

ws_data.freeze_panes = ws_data["A3"]

# ═══════════════════════════════════════════════════════════════════════════════
#  SHEET 2 — Contract Timeline  (reads from Contract Data via formulas)
# ═══════════════════════════════════════════════════════════════════════════════
ws_tl = wb.create_sheet("Contract Timeline")

# Column layout for Timeline
TC_CLIENT   = 1
TC_TRADE    = 2
TC_START    = 3
TC_END      = 4
TC_DUR      = 5
TC_GANTT_0  = 6

ws_tl.column_dimensions[get_column_letter(TC_CLIENT)].width  = 20
ws_tl.column_dimensions[get_column_letter(TC_TRADE)].width   = 22
ws_tl.column_dimensions[get_column_letter(TC_START)].width   = 13
ws_tl.column_dimensions[get_column_letter(TC_END)].width     = 13
ws_tl.column_dimensions[get_column_letter(TC_DUR)].width     = 14
for i in range(len(months)):
    ws_tl.column_dimensions[get_column_letter(TC_GANTT_0 + i)].width = 4.5

# Row 1: title
ws_tl.row_dimensions[1].height = 28
last_col = TC_GANTT_0 + len(months) - 1
ws_tl.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
t2 = ws_tl.cell(1, 1, "CONTRACT LENGTH TRACKER")
t2.font      = Font(name="Calibri", bold=True, size=16, color=C_WHITE)
t2.fill      = fill(C_NAVY)
t2.alignment = Alignment(horizontal="center", vertical="center")

# Row 2: year bands + info column headers
ws_tl.row_dimensions[2].height = 16
for c in range(1, TC_GANTT_0):
    ws_tl.cell(2, c).fill = fill(C_NAVY)

years_bounds = {}
for i, (yr, mo) in enumerate(months):
    years_bounds.setdefault(yr, [i, i])
    years_bounds[yr][1] = i

for yr, (s, e) in years_bounds.items():
    c1, c2 = TC_GANTT_0 + s, TC_GANTT_0 + e
    ws_tl.merge_cells(start_row=2, start_column=c1, end_row=2, end_column=c2)
    cell = ws_tl.cell(2, c1, str(yr))
    cell.font      = Font(name="Calibri", bold=True, size=10, color=C_WHITE)
    cell.fill      = fill(C_BLUE)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = border()

# Row 3: column headers + month abbreviations
ws_tl.row_dimensions[3].height = 20
for col, label in {TC_CLIENT: "Client", TC_TRADE: "Trade / Service",
                   TC_START: "Start Date", TC_END: "End Date",
                   TC_DUR: "Duration (days)"}.items():
    c = ws_tl.cell(3, col, label)
    c.font      = Font(name="Calibri", bold=True, size=10, color=C_WHITE)
    c.fill      = fill(C_NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = border(top=thick, bottom=thick)

for i, (yr, mo) in enumerate(months):
    col  = TC_GANTT_0 + i
    cell = ws_tl.cell(3, col, calendar.month_abbr[mo])
    cell.font      = Font(name="Calibri", bold=True, size=8, color=C_WHITE)
    cell.fill      = fill(C_NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = border(top=thick, bottom=thick)
    if (yr, mo) == (today.year, today.month):
        cell.font   = Font(name="Calibri", bold=True, size=8, color=C_TODAY)
        cell.border = Border(top=thick, bottom=thick,
                             left=Side(style="medium", color=C_TODAY),
                             right=Side(style="medium", color=C_TODAY))

# Assign a color index per client (same order as contract list)
client_color: dict[str, int] = {}
color_idx = 0
for client, *_ in contracts:
    if client not in client_color:
        client_color[client] = color_idx % len(GANTT_COLORS)
        color_idx += 1

# Data rows — formulas pull from Contract Data
for ri, (client, *_) in enumerate(contracts, 1):
    tl_row   = ri + 3        # timeline Excel row
    data_row = ri + 2        # Contract Data Excel row

    row_fill = fill(C_ALT if ri % 2 == 0 else C_WHITE)
    gantt_hex = GANTT_COLORS[client_color[client]]

    ws_tl.row_dimensions[tl_row].height = 18

    # Info columns — live references to Contract Data
    ref = lambda col_name: f"='Contract Data'!{get_column_letter(DC[col_name])}{data_row}"
    info = {
        TC_CLIENT: ref("client"),
        TC_TRADE:  ref("trade"),
        TC_START:  ref("start"),
        TC_END:    ref("end"),
        TC_DUR:    ref("dur"),
    }
    for col, formula in info.items():
        c = ws_tl.cell(tl_row, col, formula)
        c.fill      = row_fill
        c.font      = Font(name="Calibri", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = border()
    # date format for start/end display
    ws_tl.cell(tl_row, TC_START).number_format = "MM/DD/YYYY"
    ws_tl.cell(tl_row, TC_END).number_format   = "MM/DD/YYYY"
    # left-align client/trade
    ws_tl.cell(tl_row, TC_CLIENT).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws_tl.cell(tl_row, TC_TRADE).alignment  = Alignment(horizontal="left", vertical="center", indent=1)

    # Gantt cells — formula returns 1 if this month overlaps the contract, else ""
    # References Contract Data start (col E) and end (col F) for this row
    start_ref = f"'Contract Data'!${get_column_letter(DC['start'])}{data_row}"
    end_ref   = f"'Contract Data'!${get_column_letter(DC['end'])}{data_row}"

    gantt_col_start = TC_GANTT_0
    gantt_col_end   = TC_GANTT_0 + len(months) - 1

    for i, (yr, mo) in enumerate(months):
        col          = TC_GANTT_0 + i
        month_start  = f"DATE({yr},{mo},1)"
        month_end    = f"EOMONTH(DATE({yr},{mo},1),0)"
        formula      = f"=IF(AND({start_ref}<={month_end},{end_ref}>={month_start}),1,\"\")"

        c = ws_tl.cell(tl_row, col, formula)
        c.fill   = row_fill
        c.border = border()
        if (yr, mo) == (today.year, today.month):
            c.border = Border(top=thin, bottom=thin,
                              left=Side(style="medium", color=C_TODAY),
                              right=Side(style="medium", color=C_TODAY))

    # Conditional formatting: if cell = 1 → colour it with this client's colour
    gantt_range = (f"{get_column_letter(gantt_col_start)}{tl_row}:"
                   f"{get_column_letter(gantt_col_end)}{tl_row}")
    ws_tl.conditional_formatting.add(
        gantt_range,
        FormulaRule(
            formula=[f"{get_column_letter(gantt_col_start)}{tl_row}=1"],
            fill=fill(gantt_hex),
            font=Font(color=gantt_hex),  # hide the "1"
        )
    )

# Legend
legend_row = 4 + len(contracts) + 2
ws_tl.cell(legend_row, 1, "LEGEND").font      = Font(bold=True, size=10, color=C_WHITE)
ws_tl.cell(legend_row, 1).fill               = fill(C_NAVY)
ws_tl.cell(legend_row, 1).alignment          = Alignment(horizontal="center", vertical="center")

for li, (client, ci) in enumerate(client_color.items()):
    r = legend_row + 1 + li
    ws_tl.cell(r, 1, "").fill          = fill(GANTT_COLORS[ci])
    ws_tl.cell(r, 2, client).font      = Font(size=10)
    ws_tl.cell(r, 2).alignment         = Alignment(vertical="center", indent=1)

today_leg = legend_row + 1 + len(client_color)
ws_tl.cell(today_leg, 1, "").border = Border(
    left=Side(style="medium", color=C_TODAY),
    right=Side(style="medium", color=C_TODAY),
)
ws_tl.cell(today_leg, 2, "Today").font      = Font(size=10, color=C_TODAY, bold=True)
ws_tl.cell(today_leg, 2).alignment          = Alignment(vertical="center", indent=1)

ws_tl.freeze_panes = ws_tl.cell(4, TC_GANTT_0)

# ── Save ──────────────────────────────────────────────────────────────────────
out_path = "/home/user/My-First-Agent/Contract_Tracker.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
